from openpyxl import load_workbook
from cimprime import cimprime
import params

class bdmediclean:

     def __init__(self, hoja: str, otrolibro: str = None, hojadefault: bool = False) -> None:
          if otrolibro:
               self.bd = load_workbook(otrolibro, read_only=False)
               self.libroPorGuardar = otrolibro
          else:
               self.bd = load_workbook(params.LIBRODATOS, read_only=False)
               self.libroPorGuardar = params.LIBRODATOS
          self.hoja_actual = hoja
          self.hojabd = self.bd[self.hoja_actual["nombrehoja"]]
          self.maxfilas = self.contarfilas()
          cimprime(titulo="CONEXION A BD",Conexion_activa=f"{self.hojabd} - >> CONEXION ABIERTA >>")
          self.datosPorGuardar = False

     def __enter__(self) -> object:
          return self

     def get_id(self, id_inicial: str="1000") -> str:
          read_id = self.hojabd.cell(
               row=self.maxfilas, 
               column=self.hoja_actual["columnas"]["id"]["num"]
               ).value
          try:
               read_id = int(read_id)
          except:
               read_id = id_inicial
          return read_id

     def getmaxfilas(self) -> int:
          self.maxfilas = self.contarfilas()
          return self.maxfilas

     def contarfilas(self) -> int:
          filas = self.hoja_actual["filainicial"]
          while True:
               celda = self.hojabd.cell(row=filas, column=1)
               if celda.value != None:
                    filas += 1
               else:
                   return filas - 1

     def mapdatos(self, fila: int=None, columnas: list=None, idy: bool=False) -> dict:
          columnas = self.hoja_actual["columnas"].keys() if not columnas else columnas
          data = {"idy":{"dato":fila}} if idy and fila else {}
          for columna in columnas:
               data[columna] = {
                    "encabezado":self.hoja_actual["columnas"][columna]["encabezado"],
                    "dato":None if not fila else self.getDato(fila, columna)
                    }
          return data

     def eliminarContenidos(self, fila: int=None) -> None:
          filainicio = self.hoja_actual["filainicial"]
          filatermino = self.maxfilas + 1
          for fila in range(filainicio, filatermino, 1) if not fila else [fila]:
               for columna in self.hoja_actual["columnas_todas"]:
                    celda = self.hojabd.cell(row=fila, column=columna)
                    celda.value = None
          self.datosPorGuardar = True

     def eliminar(self, fila: int) -> bool:
          if type(fila) != int:
              try:
                  fila = int(fila)
              except Exception as e:
                  print(e)
                  return False
          self.hojabd.delete_rows(fila)
          self.datosPorGuardar = True
          return True

     def buscafila(self, columna: int = 1) -> int:
          filainicio = self.hoja_actual["filainicial"]
          filatermino = self.maxfilas + 2
          if type(columna) == str:
               columna = self.hoja_actual["columnas"][columna]["num"]
          for fila in range(filainicio, filatermino, 1):
               celda = self.hojabd.cell(row=fila, column=columna)
               if celda.value == None:
                    return fila

     def buscadato(self, dato: str, filainicio: int = None, columna: str = None, exacto: bool = False, filtropuntuacion: bool = False, buscartodo: bool = False) -> int:
          def buscador(filainicio: int) -> int:
               for fila in range(filainicio, maxfilas, 1):
                    celda = self.hojabd.cell(row=fila, column=self.hoja_actual["columnas"][columna]["num"])
                    try:
                         valorcelda = str(celda.value) if exacto else str(celda.value).lower()
                         datob = str(dato) if exacto else dato.lower()
                    except:
                         valorcelda = str(celda.value)
                         datob = str(dato)
                    finally:
                         if filtropuntuacion:
                              datob = datob.replace(".", "").replace(" ", "").replace("-", "")
                              valorcelda = valorcelda.replace(".", "").replace(" ", "").replace("-", "")
                         if datob == valorcelda and exacto:
                              return fila
                         elif datob in valorcelda and not exacto:
                              return fila
               else:
                    fila = None
               return fila
          
          filainicio = self.hoja_actual["filainicial"] if not filainicio else filainicio
          maxfilas = self.maxfilas + 1

          if buscartodo:
               filas = []
               while (filainicio <= maxfilas):
                    hallado = buscador(filainicio)
                    if hallado:
                         filas.append(hallado)
                    else:
                         break
                    filainicio = hallado + 1
               return filas
          else:
               return buscador(filainicio)

     def listar(self, filainicial: int = None, filas: list=None, columnas: list = None, encabezados: int = None, solodatos: bool = False, idy: bool=False) -> map:
          if not filainicial:
               filainicial = self.hoja_actual["filainicial"]
          if not columnas:
               columnas = self.hoja_actual["columnas_todas"]
          else:
               cols = []
               for columna in columnas:
                    cols.append(self.hoja_actual["columnas"][columna]["num"])
               columnas = cols
          if not encabezados:
               encabezados = self.hoja_actual["encabezados"]

          resultados = {}
          maxfilas = self.maxfilas + 1

          # Extraccion de datos "filas"
          resultados["datos"] = []
          rango = range(filainicial, maxfilas, 1) if not filas else filas
          for fila in rango:
               datafile = []
               for campo in columnas:
                    dato = self.hojabd.cell(row=fila, column=campo)
                    datafile.append(str(dato.value))
               if idy:
                    datafile.append(fila)
               resultados["datos"].append(datafile)
          if solodatos:
               return resultados["datos"]

          # Extraccion de encabezados
          resultados["encabezados"] = []
          cols = list(self.hoja_actual["columnas"].keys())
          cols.insert(0,None)
          for campo in columnas:
               encabezado = self.hoja_actual["columnas"][cols[campo]]["encabezado"]
               resultados["encabezados"].append(encabezado)
          if idy:
               resultados["encabezados"].append("idy")
          return resultados

     def putDato(self, dato: str = None, fila: int = None, columna: str = None) -> bool:
          if not fila:
               fila = self.maxfilas + 1
          celda = self.hojabd.cell(row=fila, column=self.hoja_actual["columnas"][columna]["num"])
          celda.value = dato
          self.datosPorGuardar = True
          return True

     def insertar_fila(self, filainsercion: int) -> None:
          self.hojabd.insert_rows(filainsercion)
          self.datosPorGuardar = True

     def getDato(self, fila: int = None, columna: str = None) -> bool:
          if type(fila) != list:
               fila = [fila]
          if type(columna) != list:
               columna = [columna]
          datos = []
          data = []
          for f in fila:
               for c in columna:
                    celda = self.hojabd.cell(row=f, column=self.hoja_actual["columnas"][c]["num"])
                    data.append(celda.value)
               datos.append(data)
          if len(datos) == 1:
               return datos[0] if len(datos[0]) > 1 else datos[0][0]
          else:
               return datos

     def guardar(self) -> bool:
          try:
               self.bd.save(self.libroPorGuardar)
          except Exception as e:
               cimprime(excepcion_guardado=e)
               return False
          else:
               cimprime(titulo="CONEXION A BD",estado_conexion=f"{self.hojabd} >>GUARDADO<< ")
               return True

     def cerrar(self) -> None:
          cimprime(titulo="CONEXION A BD",estado_conexion=f"{self.hojabd} >>CONEXION BD CERRADA<< ")
          self.bd.close()

     def __exit__(self, exc_type, exc_value, traceback) -> None:
          if self.datosPorGuardar:
               self.guardar()
          self.cerrar()
